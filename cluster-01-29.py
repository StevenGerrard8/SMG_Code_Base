from collections import defaultdict
import re
import pandas as pd
import numpy as np
import sys
import os
import colorsys
from openpyxl.styles import PatternFill
import argparse
import multiprocessing
from numba import njit
import time

@njit
def slope_distance_numba(row1, row2):
    """Numba-optimized slope distance calculation"""
    v = row1 - row2
    v_sorted = np.sort(v)
    n = len(v_sorted)
    distance = 0
    for k in range(n):
        distance += (2 * k - n + 1) * v_sorted[k]
    return 2.0 * distance

def generate_unique_colors(num_clusters):
    """Generate a large number of unique colors using HSV color space"""
    colors = []
    for i in range(num_clusters):
        hue = (0.618033988749895 * i) % 1.0
        saturation = 0.5 + (0.5 * (i % 2))
        value = 0.7 + (0.3 * ((i // 2) % 2))
        
        rgb = colorsys.hsv_to_rgb(hue, saturation, value)
        hex_color = ''.join([f'{int(x * 255):02X}' for x in rgb])
        
        colors.append(hex_color)
    
    return colors

def extract_numeric_suffix(name: str) -> int:
    if isinstance(name, bytes):
        name = name.decode('utf-8', errors='ignore')
    else:
        name = str(name)
    match = re.search(r'(\d+)$', name)
    if match:
        return int(match.group(1))
    return 0

def process_row(args):

    # Separate function for processing rows to enable multiprocessing

    df_slice, start, end, threshold = args
    results = []
    n=df_slice.shape[0]
    for i in range(start, end):
        row_i = df_slice[i]
        for j in range(i + 1, n):
            dist = slope_distance_numba(row_i, df_slice[j])
            results.append((i, j, dist))
    return results

def parallel_slope_distance(df, threshold):
    """Parallel processing of slope distance calculations"""
    n = df.shape[0]

    dist_matrix = np.zeros((n, n), dtype=np.float64)
    uf = UnionFind(n)

    # Chunk up the rows among CPU cores
    cpu_core_count = os.cpu_count()
    chunk_size = (n // cpu_core_count) + 1
    args_list = []
    start = 0
    for _ in range(cpu_core_count):
        end = min(start + chunk_size, n)
        if start < n:
            args_list.append((df, start, end, threshold))
        start = end

    with multiprocessing.Pool(cpu_core_count) as pool:
        all_results = pool.map(process_row, args_list)

    # Combine results, fill in dist_matrix and do union
    for results in all_results:
        for (i, j, dist) in results:
            dist_matrix[i, j] = dist
            dist_matrix[j, i] = dist
            if dist < threshold:
                uf.union(i, j)

    return dist_matrix, uf

class UnionFind:
    """Optimized UnionFind with path compression"""
    def __init__(self, n):
        self.parent = np.arange(n, dtype=np.int32)
        self.rank = np.zeros(n, dtype=np.int8)

    def find(self, x):
        """Path compression find method"""
        if self.parent[x] != x:
            self.parent[x] = self.find(self.parent[x])
        return self.parent[x]

    def union(self, x, y):
        """Union by rank"""
        root_x, root_y = self.find(x), self.find(y)
        if root_x == root_y:
            return
        
        if self.rank[root_x] < self.rank[root_y]:
            root_x, root_y = root_y, root_x
        
        self.parent[root_y] = root_x
        if self.rank[root_x] == self.rank[root_y]:
            self.rank[root_x] += 1

def check_blank(data, file_name):
    """Check for blank rows"""
    if data.iloc[:, 0].isnull().any():
        raise ValueError(f"Blank row in '{file_name}'")

def check_duplicate(data, file_name):
    """Check for duplicate entries"""
    if data.iloc[:, 0].duplicated().any():
        raise ValueError(f"Duplicate object in '{file_name}'")

def min_max_scale(x):
    """Numpy-based min-max scaling"""
    x_min, x_max = x.min(), x.max()
    return (x - x_min) / (x_max - x_min) if x_max != x_min else x * 0

def calculate_distance_matrix(df, group_order):
    """Optimize distance matrix calculation"""
    n = len(group_order)
    dist_matrix = pd.DataFrame(index=group_order, columns=group_order, dtype=float)
    data = df.iloc[:, 1:].to_numpy(dtype=np.float64) # only use slice for 1 time
    for i in range(n):
        for j in range(i, n):
            if i != j:
                distance = slope_distance_numba(
                    data[i],
                    data[j]
                )
                dist_matrix.iloc[i, j] = distance
                dist_matrix.iloc[j, i] = distance
            else:
                dist_matrix.iloc[i, j] = 0.0
    
    return dist_matrix

def save_colored_network(dist_matrix, groups, threshold, output_file, names):
    # Generate colors and sort the groups
    colors = generate_unique_colors(len(groups))
    group_list = list(groups.items())
    group_list.sort(key=lambda x: min(extract_numeric_suffix(m) for m in x[1]))
    new_order = []
    for _, members in group_list:
        members.sort(key=lambda x: extract_numeric_suffix(x)) #sort inside
        new_order.extend(members)
    dist_matrix = pd.DataFrame(dist_matrix, index=names, columns=names)
    df_dist = dist_matrix.loc[new_order, new_order]

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Convert dist_matrix to DataFrame for easier Excel writing
        df_dist.to_excel(writer, sheet_name='Distance Matrix', index=True)
        worksheet = writer.book['Distance Matrix']

        # Fill colors
        color_index = 0
        for group_key, members in groups.items():
            if len(members) >= 2:
                fill = PatternFill(
                    start_color=colors[color_index],
                    end_color=colors[color_index],
                    fill_type='solid'
                )
                color_index += 1

                for member in members:
                    row_idx = df_dist.index.get_loc(member) + 2
                    # +2 for 1-based Excel index plus header row
                    name_cell = worksheet.cell(row=row_idx, column=1)
                    name_cell.fill = fill

                    # Color the threshold connections
                    for other in df_dist.columns:
                        if other != member:
                            dist_val = df_dist.loc[member, other]
                            if dist_val < threshold:
                                col_idx = df_dist.columns.get_loc(other) + 2
                                cell = worksheet.cell(row=row_idx, column=col_idx)
                                cell.fill = fill

def main():
    # Set up argument parser
    parser = argparse.ArgumentParser(description='Optimized Cluster Analysis')
    parser.add_argument('-i', '--input', required=True, help='Input CSV file')
    parser.add_argument('-o', '--output', required=True, help='Output Excel file')
    parser.add_argument('-t', '--threshold', type=float, help='Custom threshold (optional)')
    args = parser.parse_args()

    try:
        # Load data and get number of columns
        col_num = len(pd.read_csv(args.input, header=None).columns)
        df = pd.read_csv(args.input, header=None, dtype={i: float for i in range(1, col_num)})
        
        # Set threshold automatically or use custom threshold
        threshold = args.threshold if args.threshold is not None else (col_num - 2)
        print(f"Using threshold: {threshold}")

        names = df.iloc[:, 0].values
        df_values = df.iloc[:, 1:]
        # Check data prerequisites silently
        check_blank(df, args.input)
        check_duplicate(df, args.input)

        # Apply min-max scaling
        df_values = df_values.apply(min_max_scale, axis=1)

        # calculate distance matrix and do uf in one run
        df_values = df_values.to_numpy(dtype=np.float64)
        dist_matrix, uf = parallel_slope_distance(df_values, threshold)

        n = df_values.shape[0]
        groups_dict = defaultdict(list)
        for i in range(n):
            root = uf.find(i)
            groups_dict[root].append(names[i])

        # Save the colored network
        save_colored_network(dist_matrix, groups_dict, threshold=threshold, output_file=args.output, names=names)

        # Print cluster information
        print("\nClustering Results:")
        for i, (key, group) in enumerate(groups_dict.items(), 1):
            if len(group) >= 2:
                print(f"Cluster {i}: {len(group)} members")

    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    start = time.perf_counter()
    main()
    end = time.perf_counter()
    print(f"运行时间: {end - start} 秒（perf_counter）")
